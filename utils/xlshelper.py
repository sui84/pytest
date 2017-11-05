#encoding=utf-8
from  openpyxl.reader.excel  import  load_workbook
from openpyxl.utils import get_column_letter

class XlsHelper(object):
    def __init__(self,filename= r'..\xlsx\sm.xlsx'):
        #读取excel2007文件
        self.filename = filename
        self.wb = load_workbook(filename = filename )
        #显示有多少张表
        #print  "Worksheet name(s):" , self.wb.get_sheet_names()

    def GetWorkSheet(self,sheetnum):
        #取第一张表
        sheetnames = self.wb.get_sheet_names()
        ws = self.wb.get_sheet_by_name(sheetnames[sheetnum])
        #显示表名，表行数，表列数
        print   "Work Sheet Titile:" ,ws.title
        print   "Work Sheet Rows:" ,ws.max_row
        print   "Work Sheet Cols:" ,ws.max_column
        return ws

    def GetRows(self,ws,rowstart,rowend,colstart,colend):
        records = []
        for rownum in range(rowstart,rowend+1):
            records.append(self.GetRow(ws,rownum,colstart,colend))
        return records

    def GetRow(self,ws,rownum,colstart,colend):
        record = []
        for colnum in range(colstart,colend):
            record.append(ws.cell(row=rownum,column=colnum).value)
        return record

    def WriteRows(self,ws,results,rowstart,rowend,colstart,colend):
        i=0
        for rownum in range(rowstart,rowend+1):
            self.WriteRow(ws,results[i],rownum,colstart,colend)
            i+=1
        ws.save(self.filename)

    def WriteRow(self,ws,result,rownum,colstart,colend):
        i=0
        print result
        for colnum in range(colstart,colend):
            ws.cell(row=rownum,column=colnum).value=result[i]
            i+=1

if __name__ == '__main__':
    xh=XlsHelper(filename= r'..\xlsx\sm.xlsx')
    ws = xh.GetWorkSheet(0)
    records= xh.GetRows(ws,2,16,1,10)
    print records[0]
